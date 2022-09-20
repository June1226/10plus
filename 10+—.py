# -*- coding: utf-8 -*-
# @Time  : 2022/8/23 8:34 PM
# Author : 拒绝内卷的小测试

import random
import openpyxl as op
from openpyxl.styles import Font


# 随机加减
def ten(num):
    ten_list = []

    while True:
        a = random.randint(1, 10)
        # 避免出现0+0
        b = random.randint(1, 10)
        c = random.choice("+-")
        if c == "+":
            # 控制和为10以内
            if a + b <= 10:
                ten_list.append(f"{a} + {b} = ")
        else:
            # 判断数字大小，避免出现负数
            if a < b:
                a, b = b, a
                ten_list.append(f"{a} - {b} = ")
        # 题目数量
        if len(ten_list) == num:
            break
    return ten_list


# 将公式写入excel表
def write_datas(num):
    datas = ten(num)

    # 新建excel表
    wb = op.Workbook()
    ws = wb['Sheet']

    # 把列表数据分成2列写入
    for i in range(0, len(datas), 2):
        n = i / 2
        del_datas = datas[i: i + 2]
        ws.cell(row=n + 1, column=1, value=del_datas[0])
        ws.cell(row=n + 1, column=3, value=del_datas[1])
    wb.save('10+-.xlsx')


# 设置excel格式
def set_style(num):
    wb = op.load_workbook("10+-.xlsx")
    ws = wb['Sheet']

    # 调整列宽,行高
    col = ["A", "B", "C"]
    for i in col:
        for j in range(1, num + 1):
            ws.column_dimensions[i].width = 30
            ws.row_dimensions[j].height = 40

    # 设置字体
    font = Font(name='黑体',
                size=30,
                color='FF000000',
                bold=False,
                italic=False,
                vertAlign=None,
                underline=None,
                strike=False)

    for row in ws[f'A1:C{int(num / 2)}']:
        for cell in row:
            cell.font = font

    wb.save("10+-.xlsx")

if __name__ == '__main__':
    num = int(input("请输入打印题目数量（双数）："))
    write_datas(num)
    set_style(num)